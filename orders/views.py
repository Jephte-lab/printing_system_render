# orders/views.py

from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.db import transaction
from datetime import timedelta
import io, json
import openpyxl
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa
import os
from django.conf import settings
from django.db.models import Q
from django.contrib.humanize.templatetags.humanize import intcomma
import calendar

from .models import Order
from .forms import OrderForm, OrderItemFormSet
from expenses.models import Expense
from expenses.forms import ExpenseForm

# -------------------------------
# PDF for Orders
# -------------------------------
def order_pdf(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    logo_path = os.path.join(settings.BASE_DIR, 'static/images/logo.png')
    context = {
        'order': order,
        'company_name': 'Fabulous Media',
        'company_address': 'Arua Park Plaza B4 253, Kampala, Uganda',
        'payment_note': 'Please use the the below credentials to complete payment.',
        'logo_url': logo_path,
    }
    html_string = render_to_string('orders/order_receipt.html', context)
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html_string.encode("UTF-8")), result)
    if pdf.err:
        return HttpResponse("Error generating PDF", status=500)
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="order_{order.id}.pdf"'
    return response

# -------------------------------
# Orders List with Expenses & Profit
# -------------------------------
@login_required
def orders_list(request):
    filter_type = request.GET.get('filter', '')
    selected_month = request.GET.get('month', '')
    selected_year = request.GET.get('year', '')
    search_query = request.GET.get('search', '').strip()
    today = timezone.localdate()

    # ---- Orders queryset ----
    orders_qs = Order.objects.all()
    # ---- Expenses queryset ----
    expenses_qs = Expense.objects.all()

    # --- Apply Filters ---
    def apply_filters(qs):
        if filter_type == 'today':
            qs = qs.filter(created_at__date=today)
        elif filter_type == 'this_week':
            start_week = today - timedelta(days=today.weekday())
            end_week = start_week + timedelta(days=6)
            qs = qs.filter(created_at__date__gte=start_week, created_at__date__lte=end_week)
        elif filter_type == 'last_week':
            start_last_week = today - timedelta(days=today.weekday() + 7)
            end_last_week = start_last_week + timedelta(days=6)
            qs = qs.filter(created_at__date__gte=start_last_week, created_at__date__lte=end_last_week)
        elif filter_type == 'this_year':
            qs = qs.filter(created_at__year=today.year)
        if selected_month:
            qs = qs.filter(created_at__month=int(selected_month))
        if selected_year:
            qs = qs.filter(created_at__year=int(selected_year))
        return qs

    orders_qs = apply_filters(orders_qs)
    expenses_qs = apply_filters(expenses_qs)

    # --- Search for orders ---
    if search_query:
        orders_qs = orders_qs.filter(
            Q(name__icontains=search_query) |
            Q(client__name__icontains=search_query) |
            Q(status__icontains=search_query)
        )

    orders_qs = orders_qs.order_by('-created_at')
    expenses_qs = expenses_qs.order_by('-created_at')

    # --- Totals ---
    total_revenue = sum(o.total_price for o in orders_qs)
    total_expenses = sum(e.amount for e in expenses_qs)
    profit = total_revenue - total_expenses

    # --- Month/Year choices ---
    month_choices = [(str(i).zfill(2), calendar.month_name[i]) for i in range(1, 13)]
    order_years = Order.objects.dates('created_at', 'year', order='DESC')
    year_choices = [y.year for y in order_years]

    context = {
        'orders': orders_qs,
        'expenses': expenses_qs,
        'filter_type': filter_type,
        'search_query': search_query,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'profit': profit,
        'month_choices': month_choices,
        'year_choices': year_choices,
    }

    return render(request, 'orders/orders_list.html', context)

# -------------------------------
# Order CRUD
# -------------------------------
@login_required
@transaction.atomic
def order_add(request):
    if request.method == 'POST':
        form = OrderForm(request.POST)
        formset = OrderItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            order = form.save()
            for item in formset.save(commit=False):
                item.order = order
                item.total_price = item.unit_price * item.quantity
                item.save()
            for deleted_item in formset.deleted_objects:
                deleted_item.delete()
            return redirect('orders_list')
    else:
        form = OrderForm()
        formset = OrderItemFormSet()
    return render(request, 'orders/order_form.html', {'form': form, 'formset': formset, 'title': 'Add Order'})


@login_required
@transaction.atomic
def order_edit(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if request.method == 'POST':
        form = OrderForm(request.POST, instance=order)
        formset = OrderItemFormSet(request.POST, instance=order)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect('orders_list')
    else:
        form = OrderForm(instance=order)
        formset = OrderItemFormSet(instance=order)
    return render(request, 'orders/order_form.html', {'form': form, 'formset': formset, 'title': f'Edit Order #{order.id}'})


@login_required
def order_detail(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    return render(request, 'orders/order_detail.html', {'order': order, 'items': order.items.all()})


@login_required
def order_delete(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    if request.method == 'POST':
        order.delete()
        return redirect('orders_list')


@csrf_exempt
def update_order_status(request, order_id):
    if request.method == 'POST':
        order = get_object_or_404(Order, id=order_id)
        data = json.loads(request.body)
        status = data.get('status')
        if status in ['pending', 'cleared']:
            order.status = status
            order.save()
            return JsonResponse({'success': True})
    return JsonResponse({'success': False})

# -------------------------------
# Export Orders to Excel
# -------------------------------
@login_required
def export_orders_month(request):
    month = request.GET.get('month')
    year = request.GET.get('year')
    if not month or not year:
        return HttpResponse("Month and year must be selected", status=400)

    orders = Order.objects.filter(created_at__year=int(year), created_at__month=int(month)).order_by('-created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Orders_{year}_{month}"

    headers = ['ID', 'Client', 'Order Name', 'Date', 'Total', 'Status']
    ws.append(headers)
    total_revenue = 0
    for o in orders:
        ws.append([o.id, o.client.name, o.name, o.created_at.strftime("%d %b %Y %H:%M"), float(o.total_price), o.status])
        total_revenue += float(o.total_price)

    ws.append([])
    ws.append(['', '', '', 'Total Revenue', total_revenue, ''])
    for i, column in enumerate(ws.columns, 1):
        max_len = max(len(str(cell.value)) for cell in column)
        ws.column_dimensions[get_column_letter(i)].width = max_len + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Orders_{year}_{month}.xlsx'
    wb.save(response)
    return response