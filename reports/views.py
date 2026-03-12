from django.template.loader import get_template
import calendar
import datetime
from django.db.models import Sum, Min
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
from xhtml2pdf import pisa

from orders.models import Order, OrderItem
from expenses.models import Expense
from datetime import timedelta
from django.utils import timezone


@login_required
def reports_dashboard(request):
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Only active orders
    orders = Order.objects.filter(status__in=['pending', 'cleared'])
    expenses = Expense.objects.all()

    # Apply filters
    if month and year:
        orders = orders.filter(created_at__year=int(year), created_at__month=int(month))
        expenses = expenses.filter(created_at__year=int(year), created_at__month=int(month))
    elif year:
        orders = orders.filter(created_at__year=int(year))
        expenses = expenses.filter(created_at__year=int(year))

    # Aggregate totals
    total_revenue = OrderItem.objects.filter(order__in=orders).aggregate(total=Sum('total_price'))['total'] or 0
    total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0
    profit = total_revenue - total_expenses

    # Month choices
    month_choices = [(str(i).zfill(2), calendar.month_name[i]) for i in range(1, 13)]

    # Dynamic year choices from orders and expenses
    first_order = orders.aggregate(Min('created_at'))['created_at__min']
    first_expense = expenses.aggregate(Min('created_at'))['created_at__min']
    first_year = min(
        first_order.year if first_order else datetime.date.today().year,
        first_expense.year if first_expense else datetime.date.today().year
    )
    current_year = datetime.date.today().year
    year_choices = list(range(first_year, current_year + 1))

    context = {
        'orders': orders,
        'expenses': expenses,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'profit': profit,
        'month': month,
        'year': year,
        'month_choices': month_choices,
        'year_choices': year_choices,
    }

    return render(request, 'reports/reports_dashboard.html', context)


@login_required
def export_reports_excel(request):
    import calendar
    from openpyxl.styles import Font, Alignment, numbers

    # --- FILTERS FROM PAGE ---
    month = request.GET.get('month')
    year = request.GET.get('year')
    filter_type = request.GET.get('filter')  # optional if you want same as dashboard

    today = timezone.localdate()

    # --- ORDERS: only active orders like on page ---
    orders_qs = Order.objects.filter(status__in=['pending', 'cleared'])
    if filter_type == 'today':
        orders_qs = orders_qs.filter(created_at__date=today)
    elif filter_type == 'this_week':
        start_week = today - timedelta(days=today.weekday())
        end_week = start_week + timedelta(days=6)
        orders_qs = orders_qs.filter(created_at__date__gte=start_week, created_at__date__lte=end_week)
    elif filter_type == 'last_week':
        start_last = today - timedelta(days=today.weekday() + 7)
        end_last = start_last + timedelta(days=6)
        orders_qs = orders_qs.filter(created_at__date__gte=start_last, created_at__date__lte=end_last)
    elif filter_type == 'this_year':
        orders_qs = orders_qs.filter(created_at__year=today.year)

    if month:
        orders_qs = orders_qs.filter(created_at__month=int(month))
    if year:
        orders_qs = orders_qs.filter(created_at__year=int(year))

    # --- EXPENSES: same filters as page ---
    expenses_qs = Expense.objects.all()
    if month:
        expenses_qs = expenses_qs.filter(created_at__month=int(month))
    if year:
        expenses_qs = expenses_qs.filter(created_at__year=int(year))

    # --- AGGREGATES ---
    total_revenue = OrderItem.objects.filter(order__in=orders_qs).aggregate(total=Sum('total_price'))['total'] or 0
    total_expenses = expenses_qs.aggregate(total=Sum('amount'))['total'] or 0
    profit = total_revenue - total_expenses

    # --- EXCEL WORKBOOK ---
    wb = openpyxl.Workbook()
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center')

    # ===== SUMMARY =====
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Metric", "Amount (UGX)"])
    for cell in ws_summary[1]:
        cell.font = bold_font
        cell.alignment = center_align
    summary_data = [["Total Revenue", total_revenue],
                    ["Total Expenses", total_expenses],
                    ["Profit", profit]]
    for row in summary_data:
        ws_summary.append([row[0], float(row[1])])
        ws_summary[f'B{ws_summary.max_row}'].number_format = '#,##0.00'
    for column_cells in ws_summary.columns:
        ws_summary.column_dimensions[column_cells[0].column_letter].width = max(len(str(c.value)) for c in column_cells) + 5

    # ===== ORDERS =====
    ws_orders = wb.create_sheet("Orders")
    ws_orders.append(["Order ID", "Client", "Order Name", "Service", "Quantity", "Unit Price (UGX)", "Total Price (UGX)", "Date"])
    for cell in ws_orders[1]:
        cell.font = bold_font
        cell.alignment = center_align
    for item in OrderItem.objects.filter(order__in=orders_qs):
        ws_orders.append([
            item.order.id,
            item.order.client.name,
            item.order.name,
            item.service.name,
            item.quantity,
            float(item.unit_price),
            float(item.total_price),
            item.order.created_at.strftime("%d %b %Y %H:%M")
        ])
        row_num = ws_orders.max_row
        ws_orders[f'F{row_num}'].number_format = '#,##0.00'
        ws_orders[f'G{row_num}'].number_format = '#,##0.00'
        ws_orders[f'H{row_num}'].alignment = center_align
    for column_cells in ws_orders.columns:
        ws_orders.column_dimensions[column_cells[0].column_letter].width = max(len(str(c.value)) for c in column_cells) + 5

    # ===== EXPENSES =====
    ws_expenses = wb.create_sheet("Expenses")
    ws_expenses.append(["Expense ID", "Title", "Note", "Amount (UGX)", "Date"])
    for cell in ws_expenses[1]:
        cell.font = bold_font
        cell.alignment = center_align
    for e in expenses_qs:
        ws_expenses.append([
            e.id,
            e.title,
            getattr(e, 'note', ''),
            float(e.amount),
            e.created_at.strftime("%d %b %Y %H:%M")
        ])
        row_num = ws_expenses.max_row
        ws_expenses[f'D{row_num}'].number_format = '#,##0.00'
        ws_expenses[f'E{row_num}'].alignment = center_align
    for column_cells in ws_expenses.columns:
        ws_expenses.column_dimensions[column_cells[0].column_letter].width = max(len(str(c.value)) for c in column_cells) + 5

    # ===== FILE NAME DYNAMIC =====
    file_name = "Financial_Report"
    if month and year:
        file_name += f"_{calendar.month_name[int(month)]}_{year}"
    elif year:
        file_name += f"_{year}"
    file_name += ".xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={file_name}'
    wb.save(response)
    return response


@login_required
def reports_pdf(request):
    month = request.GET.get('month')
    year = request.GET.get('year')

    orders = Order.objects.all()
    expenses = Expense.objects.all()

    if month and year:
        orders = orders.filter(created_at__year=int(year), created_at__month=int(month))
        expenses = expenses.filter(created_at__year=int(year), created_at__month=int(month))
    elif year:
        orders = orders.filter(created_at__year=int(year))
        expenses = expenses.filter(created_at__year=int(year))

    order_items = OrderItem.objects.filter(order__in=orders)

    total_revenue = order_items.aggregate(total=Sum('total_price'))['total'] or 0
    total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0
    profit = total_revenue - total_expenses

    # Format numbers BEFORE template (PDF-safe)
    total_revenue_str = f"{total_revenue:,.2f}"
    total_expenses_str = f"{total_expenses:,.2f}"
    profit_str = f"{profit:,.2f}"

    month_name = calendar.month_name[int(month)] if month else None

    template = get_template("reports/reports_pdf.html")

    context = {
        "orders": orders,
        "order_items": order_items,
        "expenses": expenses,
        "total_revenue": total_revenue_str,
        "total_expenses": total_expenses_str,
        "profit": profit_str,
        "month_name": month_name,
        "year": year,
        "profit_negative": profit < 0,
    }

    html = template.render(context)

    response = HttpResponse(content_type="application/pdf")
    filename = "Financial_Report"
    if month_name:
        filename += f"_{month_name}"
    if year:
        filename += f"_{year}"
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("Error generating PDF", status=500)

    return response